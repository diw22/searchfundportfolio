import tkinter as tk
from tkinter import ttk, filedialog, messagebox, Text
import pandas as pd
import sqlite3
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from googletrans import Translator, LANGUAGES
from datetime import datetime

##linkedin FTEs
#translation saved and used
#keyword exclusion
#edge cases


rejected_companies =[]

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

country_language_map = {
    'Switzerland': 'fr',  
    'France': 'fr',       
    'Spain': 'es',               
    'Denmark': 'da',      
    'Netherlands': 'nl', 
    'Portugal': 'pt'      
}

country_currency_map = {
    'Switzerland': 'CHF',
    'France': 'EUR',
    'Spain': 'EUR',
    'Germany': 'EUR',
    'Italy': 'EUR',
    'Denmark': 'DKK',
    'Netherlands': 'EUR',
    'Portugal': 'EUR',
    'United Kingdom': 'GBP',
    'UK': 'GBP'
}

countries_of_interest = ['UK', 'United Kingdom', 'Spain', 'Portugal', 'France', 'Switzerland', 'Netherlands', 'Denmark']

translation_cache = {}

# prep
def extract_country(location):
    if pd.isna(location):
        return None
    if ',' in location:
        parts = location.split(', ')
        if len(parts) > 1:
            return parts[-1]
    else:
        return location
    return None

def map_currency(country):
    return country_currency_map.get(country, None)



def extract_city(location):
    if pd.isna(location):
        return None
    if ',' in location:
        parts = location.split(', ')
        if len(parts) > 1:
            return parts [0] 
    else:
        return location
    return None

# check
def check_website_for_keywords(url, keywords_set_1, keywords_set_2):
    try:
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
        response = requests.get(url, verify=False, timeout=10)
        if response.status_code == 200:
            content = response.text.lower()
            found_in_set_1 = any(keyword in content for keyword in keywords_set_1)
            found_in_set_2 = any(keyword in content for keyword in keywords_set_2)
            return found_in_set_1 and found_in_set_2
    except requests.RequestException as e:
        print(f"Error accessing {url}")
    return False

def keyword_and_country_filter(company_name, description, inclusion_keywords_1, inclusion_keywords_2, location, countries_of_interest, website, FTE, min_FTE=10):
    country = extract_country(location)
    found_in_set_1 = []
    found_in_set_2 = []
    found_in_translated_set_1 = []
    found_in_translated_set_2 = []
    found_in_translated_set_1_bool = False
    found_in_translated_set_2_bool = False
    if country and country in countries_of_interest and FTE >= min_FTE:
        description_lower = description.lower() if description else ''
        
    
        processed_keywords_1 = [keyword.strip() for keyword in ','.join(inclusion_keywords_1).split(',')]
        processed_keywords_2 = [keyword.strip() for keyword in ','.join(inclusion_keywords_2).split(',')]
      
        found_in_set_1_bool = any(keyword in description_lower for keyword in processed_keywords_1)
        found_in_set_2_bool = any(keyword in description_lower for keyword in processed_keywords_2)
        found_in_set_1 = [keyword for keyword in processed_keywords_1 if keyword in description_lower]
        found_in_set_2 = [keyword for keyword in processed_keywords_2 if keyword in description_lower]
        if found_in_set_1_bool and found_in_set_2_bool:
            return True, country, website, True
        
        website_has_keywords = check_website_for_keywords(website, processed_keywords_1, processed_keywords_2)
        
        if website_has_keywords:
            return True, country, website, True

       
       
        if country in country_language_map:
            local_language = country_language_map[country]
            translated_keywords_1 = translate_keywords(inclusion_keywords_1, local_language)
            translated_keywords_2 = translate_keywords(inclusion_keywords_2, local_language)
            processed_translated_keywords_1 = [keyword.strip() for keyword in ','.join(translated_keywords_1).split(',')]
            processed_translated_keywords_2 =[keyword.strip() for keyword in ','.join(translated_keywords_2).split(',')]
            found_in_translated_set_1_bool = any(keyword in description_lower for keyword in processed_translated_keywords_1)
            found_in_translated_set_2_bool = any(keyword in description_lower for keyword in processed_translated_keywords_2)
            found_in_translated_set_1 = [keyword for keyword in processed_translated_keywords_1 if keyword in description_lower]
            found_in_translated_set_2 = [keyword for keyword in processed_translated_keywords_2 if keyword in description_lower]
            if found_in_translated_set_1_bool and found_in_translated_set_2_bool:
                return True, country, website, True
            
            website_has_translated_keywords = check_website_for_keywords(website, processed_translated_keywords_1, processed_translated_keywords_2)
            
            if website_has_translated_keywords:
                return True, country, website, True
            
    
        reason = []
        if not found_in_set_1_bool:
            reason.append("Missing keywords from set 1")
        if not found_in_set_2_bool:
            reason.append("Missing keywords from set 2")
        if not found_in_translated_set_1_bool:
            reason.append("Missing translated keywords from set 1")
        if not found_in_translated_set_2_bool:
            reason.append("Missing translated keywords from set 2")
        if FTE < min_FTE:
            reason.append("FTE below minimum")
        rejected_companies.append({
            "Company": company_name,
            "Description": description,
            "Reason": ", ".join(reason),
            "Product Keywords Found": found_in_set_1,
            "Subvertical Keywords Found": found_in_set_2,
            "Translated Product Keywords Found": found_in_translated_set_1,
            "Translated Subvertical Keywords Found": found_in_translated_set_2
        })
    return False, country, website, False

def translate_keywords():
    translator = Translator()
    keywords = keywords_textbox.get("1.0", tk.END).strip().split('\n')
    translated_keywords = {}

    for keyword in keywords:
        for country, lang_code in country_language_map.items():
            if lang_code not in translated_keywords:
                translated_keywords[lang_code] = []
            translated_keyword = translator.translate(keyword, dest=lang_code).text
            translated_keywords[lang_code].append(translated_keyword)
    
    # Display translated keywords or use them as needed
    for lang_code, translations in translated_keywords.items():
        print(f"Translations in {LANGUAGES[lang_code]}: {', '.join(translations)}")

#def translate_keywords(keywords, language):
#    global translation_cache
#    translator = Translator()
#    translated_keywords = []
#
#    for keyword in keywords:
#        cache_key = (keyword, language)
#        if cache_key in translation_cache:
#            translated_keywords.append(translation_cache[cache_key])
#        else:
#            try:
#                translated = translator.translate(keyword, dest=language)
#                if translated and translated.text:
#                    translated_text = translated.text
#                    translation_cache[cache_key] = translated_text
#                    translated_keywords.append(translated_text)
#                    print(translated_keywords)
#            except Exception as e:
#                print(f"Translation error for '{keyword}', skipping to the next word")
#                continue
#
#    return translated_keywords

def open_language_selection_screen(keywords_textbox):
    
    language_window = tk.Toplevel(root)
    language_window.title("Select Language")
    keywords = keywords_textbox.get(1.0, tk.END).strip().split("\n")
    
    def open_translation_screen(lang_code, keywords_1):
        
        translations = translate_keywords(keywords_1, lang_code)
        open_translation_textbox(translations)

    for country, lang in country_language_map.items():
        button = tk.Button(language_window, text=country, command=lambda l=lang: open_translation_screen(l, keywords))
        button.pack()
        
def open_translation_textbox(translations):
    
    translation_window = tk.Toplevel(root)
    translation_window.title("Edit Translations")

    text_box = tk.Text(translation_window, wrap='word')
    text_box.pack(expand=True, fill='both')

    for translation in translations:
        text_box.insert('end', translation + '\n')
        
    save_button_1 = tk.Button(translation_window, text="Save", command=lambda:save_translations)
    save_button_1.pack()
    
    def save_translations():
            file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
            if file_path:
                with open(file_path, 'w') as file:
                    file.write(text_box.get('1.0', 'end'))


def filter_function(db_path, excel_path, keywords_list_1, keywords_list_2, title_1, title_2, title_3, username):
    output_csv_path = db_path
    output_df = pd.read_csv(output_csv_path)
    total_items = len(output_df)


    progress["value"] = 0
    progress["maximum"] = total_items

    #add name thing
    included_company_names = []
    included_countries = []
    included_websites = []
    included_groups = []
    included_Linkedins = []
    included_foundingyears = []
    date_added = datetime.now().strftime('%d/%m/%y')
    added_by = username
    included_cities = []
    included_currencies = []
    included_verticals = []
    included_subverticals = []
    included_FTEs = []

    for i, row in enumerate(output_df.iterrows()):
        _, row = row
        description = row['Description']
        location = row.get('Headquarters', row.get('Country', row.get('Address')))
        company_name = row.get('Name', row.get('Company Name'))
        website = row.get('Domain', row.get('Website'))
        founding_year = row.get('Year Founded')
        linkedin = row.get('LinkedIn')
        city = extract_city(location)
        currency = map_currency(extract_country(location))
        try:
            FTE = int(row.get('Employee Estimate', 0))
        except ValueError:
            FTE = 0

        is_included, country, website, is_software = keyword_and_country_filter(company_name, description, keywords_list_1, keywords_list_2, location, countries_of_interest, website, FTE, min_FTE=10)
        if is_included:
            included_company_names.append(company_name)
            included_countries.append(country)
            included_websites.append(website)
            included_groups.append(title_1 if is_software else "")
            included_verticals.append(title_3 if is_included else "")
            included_subverticals.append(title_2 if is_included else "")
            included_Linkedins.append(linkedin)
            included_foundingyears.append(founding_year)
            included_cities.append(city)
            included_currencies.append(currency)
            included_FTEs.append(FTE)

       
        if i % 10 == 0 or i == total_items - 1:
            progress["value"] = i + 1
            root.update_idletasks()

    output_path = excel_path
    workbook = load_workbook(output_path)
    worksheet = workbook['2. Qualified Leads']

    first_empty_row = None
    for row in worksheet.iter_rows(min_row=9, min_col=4, max_col=4):
        if not row[0].value:
            first_empty_row = row[0].row
            break

    if first_empty_row is None:
        first_empty_row = worksheet.max_row + 1

    for i, (company_name, country, website, group, founding_year, linkedin, city, currency, vertical, subvertical, FTE) in enumerate(zip(included_company_names, included_countries, included_websites, included_groups, included_foundingyears, included_Linkedins, included_cities, included_currencies, included_verticals, included_subverticals, included_FTEs), start=first_empty_row):
        worksheet.cell(row=i, column=4, value=company_name)
        worksheet.cell(row=i, column=5, value=country)
        worksheet.cell(row=i, column=6, value=website)
        worksheet.cell(row=i, column=7, value=group)
        worksheet.cell(row=i, column=8, value=vertical)
        worksheet.cell(row=i, column=9, value=subvertical)
        worksheet.cell(row=i, column=10, value=date_added)
        worksheet.cell(row=i, column=11, value=added_by)
        worksheet.cell(row=i, column=13, value='Proprietary')
        worksheet.cell(row=i, column=14, value='Target')
        worksheet.cell(row=i, column=15, value=founding_year)
        worksheet.cell(row=i, column=16, value=city)
        worksheet.cell(row=i, column=17, value=linkedin)
        worksheet.cell(row=i, column=18, value=currency)
        worksheet.cell(row=i, column=20, value=FTE)

    workbook.save(excel_path)

def qualify_function():

    pass

def open_file_selector(entry_widget):
    filename = filedialog.askopenfilename()
    entry_widget.delete(0, tk.END)
    entry_widget.insert(0, filename)

def clear_inputs():
    db_entry.delete(0, tk.END)
    excel_entry.delete(0, tk.END)
    subvertical_entry.delete(0, tk.END)
    product_entry.delete(0, tk.END)
    subvertical_keywords_textbox.delete(1.0, tk.END)
    keyword_textbox.delete(1.0, tk.END)
    user_entry.delete(0, tk.END)
    vertical_entry.delete(0, tk.END)

def filter_action():
    db_path = db_entry.get()
    excel_path = excel_entry.get()
    keywords_list_1 = keyword_textbox.get(1.0, tk.END).strip().split("\n")
    keywords_list_2 = subvertical_keywords_textbox.get(1.0, tk.END).strip().split("\n")
    title_1 = product_entry.get()
    title_2 = subvertical_entry.get()
    title_3 = vertical_entry.get()
    username = user_entry.get()
    filter_function(db_path, excel_path, keywords_list_1, keywords_list_2, title_1, title_2, title_3, username)
    messagebox.showinfo("Info", "Filter action completed")

def qualify_action():
    qualify_function()
    messagebox.showinfo("Info", "Qualify action completed")


def browse_and_load_keywords(keywords_text):
    
    filename = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
    if filename:
        with open(filename, 'r') as file:
            content = file.read()
            keywords_text.delete(1.0, tk.END)  
            keywords_text.insert(tk.END, content)  

def save_keywords_to_file(keywordstext):
    try:
        keywords = keywordstext.get("1.0", "end-1c")  
        if not keywords.strip():
            messagebox.showwarning("No Keywords", "Please enter some keywords before saving.")
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt"), ("All files", "*.*")])
        if file_path:
            with open(file_path, 'w') as file:
                file.write(keywords)
            messagebox.showinfo("Success", "Keywords saved successfully.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while saving the file: {e}")

def show_translated_keywords(self, language):
        if not hasattr(self, 'translations'):
            messagebox.showwarning("Warning", "Please load the keywords first.")
            return

        new_window = tk.Toplevel(self)
        new_window.title(f"{language} Keywords")
        
        text = tk.Text(new_window, width=50, height=20)
        text.pack(padx=10, pady=10)
        
        translated_keywords = ', '.join(self.translations[language])
        text.insert(tk.END, translated_keywords)
        
        save_button = tk.Button(new_window, text="Save", 
                                command= lambda:self.save_translated_keywords(language, text.get("1.0", tk.END)))
        save_button.pack(pady=5)

#        
def save_translated_keywords(self, language, keywords):
        self.translations[language] = keywords.strip().split(', ')
        messagebox.showinfo("Info", f"{language} keywords updated and saved.")

def export_rejected_companies_to_excel(reject_pile):
    reject_path = reject_pile.get()
    df = pd.DataFrame(rejected_companies)
    df.to_excel(reject_path, index=False)

root = tk.Tk()
root.title("Database Filter")
root.configure(bg='light blue')

def browsefunc(entry):
    filename = filedialog.askopenfilename()
    entry.delete(0, tk.END)
    entry.insert(tk.END, filename)


# database path
db_label = tk.Label(root, text="Database Path:", bg='light blue', fg='dark blue')
db_label.grid(row=0, column=0)
db_entry = tk.Entry(root, width=50)
db_entry.grid(row=0, column=1)
db_browse = tk.Button(root, text="Browse", command=lambda: browsefunc(db_entry), bg='dark blue', fg='white')
db_browse.grid(row=0, column=2)

# excel path
excel_label = tk.Label(root, text="Excel Path:", bg='light blue', fg='dark blue')
excel_label.grid(row=1, column=0)
excel_entry = tk.Entry(root, width=50)
excel_entry.grid(row=1, column=1)
excel_browse = tk.Button(root, text="Browse", command=lambda: browsefunc(excel_entry), bg='dark blue', fg='white')
excel_browse.grid(row=1, column=2)

#reject path
reject_label = tk.Label(root, text="Reject Path:", bg='light blue', fg='dark blue')
reject_label.grid(row=2, column=0)
reject_entry = tk.Entry(root, width=50)
reject_entry.grid(row=2, column=1)
reject_browse = tk.Button(root, text="Browse", command=lambda: browsefunc(reject_entry), bg='dark blue', fg='white')
reject_browse.grid(row=2, column=2)

# User Name
user_label = tk.Label(root, text="User Name:", bg='light blue', fg='dark blue')
user_label.grid(row=3, column=0)
user_entry = tk.Entry(root)
user_entry.grid(row=3, column=1)

# product
product_label = tk.Label(root, text="Product:", bg='light blue', fg='dark blue')
product_label.grid(row=4, column=0)
product_entry = tk.Entry(root)
product_entry.grid(row=4, column=1)

# product keywords
keywords_label = tk.Label(root, text="Keywords:", bg='light blue', fg='dark blue')
keywords_label.grid(row=5, column=0)
keyword_textbox = Text(root, height=10, width=50)
keyword_textbox.grid(row=5, column=1)

# product preset
browse_button = tk.Button(root, text="Browse and Load Keywords", command= lambda:browse_and_load_keywords(keyword_textbox), bg='light blue', fg='dark blue')
browse_button.grid(row=5, column=4)
translate_keywords_button = tk.Button(root, text="Translate Keywords", command= lambda:translate_keywords)
#load_keywords_button = tk.Button(root, text="Translate Keywords", command= lambda:open_language_selection_screen(keyword_textbox), bg='dark blue', fg='white')
translate_keywords_button.grid(row=5, column=6)

# subvertical
subvertical_label = tk.Label(root, text="Subvertical:", bg='light blue', fg='dark blue')
subvertical_label.grid(row=6, column=0)
subvertical_entry = tk.Entry(root)
subvertical_entry.grid(row=6, column=1)

# subvertical keywords
subvertical_keywords_label = tk.Label(root, text="Keywords:", bg='light blue', fg='dark blue')
subvertical_keywords_label.grid(row=7, column=0)
subvertical_keywords_textbox = Text(root, height=10, width=50)
subvertical_keywords_textbox.grid(row=7, column=1)

# subvertical presets
browse_button_2 = tk.Button(root, text="Browse and Load Keywords", command= lambda:browse_and_load_keywords(subvertical_keywords_textbox), bg='light blue', fg='dark blue')
browse_button_2.grid(row=7, column=4)

load_keywords_button_2 = tk.Button(root, text="Translate Keywords", command= lambda:open_language_selection_screen(subvertical_keywords_textbox), bg='dark blue', fg='white')
load_keywords_button_2.grid(row=7, column=6)

# vertical
vertical_label = tk.Label(root, text="Vertical:", bg='light blue', fg='dark blue')
vertical_label.grid(row=8, column=0)
vertical_entry = tk.Entry(root)
vertical_entry.grid(row=8, column=1)

# progress bar
progress = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
progress.grid(row=9, column=0, columnspan=3, pady=10)

save_button = tk.Button(root, text="Save Keywords", command=lambda:save_keywords_to_file(keyword_textbox))
save_button.grid(row=5, column=5)

save_button_2 = tk.Button(root, text="Save Keywords", command=lambda:save_keywords_to_file(subvertical_keywords_textbox))
save_button_2.grid(row=7, column=5)


tk.Button(root, text="Clear", command=clear_inputs, bg='dark blue', fg='white').grid(row=11, column=0)
tk.Button(root, text="Filter", command=filter_action, bg='dark blue', fg='white').grid(row=11, column=1)
#tk.Button(root, text="Print rejected companies", command= export_rejected_companies_to_excel(reject_entry), bg='dark blue', fg='white').grid(row=11, column=2)
root.mainloop()
